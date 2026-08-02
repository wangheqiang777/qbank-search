import openpyxl

PATH = r"C:\Users\Administrator\Desktop\新2025强基练兵题库单选、多选、判断.xlsx"

wb = openpyxl.load_workbook(PATH, read_only=True, data_only=True)
print("SHEETS:", wb.sheetnames)

for ws in wb.worksheets:
    print("\n" + "=" * 70)
    print(f"SHEET: {ws.title} | max_row={ws.max_row} max_col={ws.max_column}")
    print("=" * 70)
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i > 5:
            break
        cells = []
        for c in row:
            s = "" if c is None else str(c)
            cells.append(s[:50])
        print(f"r{i}: {cells}")
